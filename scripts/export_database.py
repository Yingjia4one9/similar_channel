"""将数据库内容导出为表格文件"""
import sqlite3
import json
import csv
import sys
from pathlib import Path
from datetime import datetime

# 修复Windows控制台编码问题
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

# 数据库路径
DB_PATH = Path(__file__).parent.parent / "data" / "channel_index.db"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "exports"
OUTPUT_DIR.mkdir(exist_ok=True)

def export_to_csv():
    """导出为CSV格式"""
    if not DB_PATH.exists():
        print(f"数据库文件不存在: {DB_PATH}")
        return
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = OUTPUT_DIR / f"channels_export_{timestamp}.csv"
    
    print("=" * 80)
    print("导出数据库内容为CSV")
    print("=" * 80)
    
    # 获取所有频道数据
    cur.execute("""
        SELECT 
            channel_id,
            title,
            description,
            subscriber_count,
            view_count,
            country,
            language,
            emails,
            topics,
            audience,
            updated_at
        FROM channels
        ORDER BY subscriber_count DESC
    """)
    
    channels = cur.fetchall()
    
    if not channels:
        print("数据库中没有数据")
        conn.close()
        return
    
    # 写入CSV文件
    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        # 写入表头
        writer.writerow([
            '频道ID',
            '标题',
            '描述',
            '订阅数',
            '观看数',
            '国家',
            '语言',
            '邮箱',
            'Topics',
            'Audience',
            '更新时间'
        ])
        
        # 写入数据
        for ch in channels:
            # 解析JSON字段
            emails = ''
            if ch['emails']:
                try:
                    emails_list = json.loads(ch['emails'])
                    emails = ', '.join(emails_list) if isinstance(emails_list, list) else str(emails_list)
                except:
                    emails = ch['emails']
            
            topics = ''
            if ch['topics']:
                try:
                    topics_list = json.loads(ch['topics'])
                    topics = ', '.join(topics_list) if isinstance(topics_list, list) else str(topics_list)
                except:
                    topics = ch['topics']
            
            audience = ''
            if ch['audience']:
                try:
                    audience_list = json.loads(ch['audience'])
                    audience = ', '.join(audience_list) if isinstance(audience_list, list) else str(audience_list)
                except:
                    audience = ch['audience']
            
            writer.writerow([
                ch['channel_id'],
                ch['title'] or '',
                ch['description'] or '',
                ch['subscriber_count'] or 0,
                ch['view_count'] or 0,
                ch['country'] or '',
                ch['language'] or '',
                emails,
                topics,
                audience,
                ch['updated_at'] or ''
            ])
    
    conn.close()
    
    print(f"✓ 成功导出 {len(channels)} 条记录")
    print(f"✓ 文件保存位置: {csv_file}")
    print(f"✓ 文件大小: {csv_file.stat().st_size / 1024:.2f} KB")
    print("\n提示: CSV文件使用UTF-8-BOM编码，可以在Excel中正确显示中文")

def export_to_excel():
    """导出为Excel格式（需要openpyxl库）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("错误: 需要安装 openpyxl 库才能导出Excel格式")
        print("请运行: pip install openpyxl")
        return False
    
    if not DB_PATH.exists():
        print(f"数据库文件不存在: {DB_PATH}")
        return False
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = OUTPUT_DIR / f"channels_export_{timestamp}.xlsx"
    
    print("=" * 80)
    print("导出数据库内容为Excel")
    print("=" * 80)
    
    # 获取所有频道数据
    cur.execute("""
        SELECT 
            channel_id,
            title,
            description,
            subscriber_count,
            view_count,
            country,
            language,
            emails,
            topics,
            audience,
            updated_at
        FROM channels
        ORDER BY subscriber_count DESC
    """)
    
    channels = cur.fetchall()
    
    if not channels:
        print("数据库中没有数据")
        conn.close()
        return False
    
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "频道数据"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 写入表头
    headers = [
        '频道ID',
        '标题',
        '描述',
        '订阅数',
        '观看数',
        '国家',
        '语言',
        '邮箱',
        'Topics',
        'Audience',
        '更新时间'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for row_idx, ch in enumerate(channels, 2):
        # 解析JSON字段
        emails = ''
        if ch['emails']:
            try:
                emails_list = json.loads(ch['emails'])
                emails = ', '.join(emails_list) if isinstance(emails_list, list) else str(emails_list)
            except:
                emails = ch['emails']
        
        topics = ''
        if ch['topics']:
            try:
                topics_list = json.loads(ch['topics'])
                topics = ', '.join(topics_list) if isinstance(topics_list, list) else str(topics_list)
            except:
                topics = ch['topics']
        
        audience = ''
        if ch['audience']:
            try:
                audience_list = json.loads(ch['audience'])
                audience = ', '.join(audience_list) if isinstance(audience_list, list) else str(audience_list)
            except:
                audience = ch['audience']
        
        ws.cell(row=row_idx, column=1, value=ch['channel_id'])
        ws.cell(row=row_idx, column=2, value=ch['title'] or '')
        ws.cell(row=row_idx, column=3, value=ch['description'] or '')
        ws.cell(row=row_idx, column=4, value=ch['subscriber_count'] or 0)
        ws.cell(row=row_idx, column=5, value=ch['view_count'] or 0)
        ws.cell(row=row_idx, column=6, value=ch['country'] or '')
        ws.cell(row=row_idx, column=7, value=ch['language'] or '')
        ws.cell(row=row_idx, column=8, value=emails)
        ws.cell(row=row_idx, column=9, value=topics)
        ws.cell(row=row_idx, column=10, value=audience)
        ws.cell(row=row_idx, column=11, value=ch['updated_at'] or '')
    
    # 调整列宽
    column_widths = [20, 30, 50, 12, 15, 8, 8, 30, 40, 40, 20]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(excel_file)
    conn.close()
    
    print(f"✓ 成功导出 {len(channels)} 条记录")
    print(f"✓ 文件保存位置: {excel_file}")
    print(f"✓ 文件大小: {excel_file.stat().st_size / 1024:.2f} KB")
    return True

def main():
    """主函数"""
    import sys
    
    # 支持命令行参数
    if len(sys.argv) > 1:
        choice = sys.argv[1]
    else:
        print("数据库导出工具")
        print("=" * 80)
        print("\n请选择导出格式:")
        print("1. CSV格式 (推荐，无需额外依赖)")
        print("2. Excel格式 (需要openpyxl库)")
        print("3. 两种格式都导出")
        print("0. 取消")
        
        try:
            choice = input("\n请输入选项 (1/2/3/0): ").strip()
        except (EOFError, KeyboardInterrupt):
            # 非交互式环境，默认导出CSV
            print("\n非交互式环境，默认导出CSV格式")
            choice = '1'
    
    if choice == '1' or choice == 'csv':
        export_to_csv()
    elif choice == '2' or choice == 'excel':
        export_to_excel()
    elif choice == '3' or choice == 'all':
        export_to_csv()
        print()
        export_to_excel()
    elif choice == '0' or choice == 'cancel':
        print("已取消")
    else:
        print("无效选项，默认导出CSV格式")
        export_to_csv()

if __name__ == "__main__":
    main()

